from selenium import webdriver 
from fake_useragent import UserAgent
import time 
import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook
from tkinter import *
from tkinter import messagebox
import tkinter as tk
import tkinter.font as tkFont
import tkinter.ttk
from tkinter import filedialog
import logging
from openpyxl.styles import PatternFill
from openpyxl.styles import Font
import pyautogui as pg

#엑셀복사
import os
import shutil
import datetime

logger = logging.getLogger("logging")
#크롬 드라이버 설치된 위치 (원화표시 안됨. '/' 로 구분할 것)
#작업사이 대기시간
SLEEPTIME = 1
#환경설정 메뉴
drvPath = 'C:/chromeDriver/chromedriver.exe'


urlWmp = 'https://front.wemakeprice.com/user/login'
url_11st  ='https://login.11st.co.kr/auth/front/login.tmall'
urlGlob = urlWmp
whatTo = "위메프"
#저녁변수, 
######################파일명 바뀌면 xlDone, xlRow 초기화하기
#filePath = "C:/Users/lasto/OneDrive/바탕 화면/원기/코딩작업/job/test01.xlsx"

########################나중에 삭제하고 입력값으로 받기
#엑셀작업 완료여부(파일이 새로 로드되면 False로 변경하기)
xlDone = False
dicFlag = False
window = Tk()
monitor_height = window.winfo_screenheight()
monitor_width = window.winfo_screenwidth() 
zipFlag = False
#글씨체
fontStyle = tkFont.Font(family="Lucida Grande", size=20)
fontStyle2 = tkFont.Font(family="Lucida Grande", size=17)

idDic = {}
#*******동적변수*********************
#파일명 동적변수
filePath = ""
fileName = ""
fileNameVar = tk.StringVar()
fileNameVar.set(fileName)

#프로그레스 동적변수
progress = ""
progressVar = tk.StringVar()
progressVar.set(progress)

#엑셀 행 동적변수
xlRow = 3
rowVar = tk.IntVar()
rowVar.set(xlRow)

#수취인명 동적변수
custNm = ""
custNmVar = tk.StringVar()
custNmVar.set(custNm)

#위메프 아이디 동적변수
wmpId = "ID"
wmpIdVar = tk.StringVar()
wmpIdVar.set(wmpId)

#상품명 동적변수
goodsNm = ""
goodsNmVar = tk.StringVar()
goodsNmVar.set(goodsNm)

#엑셀 행 동적변수
goodsCnt = ""
goodsCntVar = tk.StringVar()
goodsCntVar.set(goodsCnt)


#쿠팡수취인 우편번호 동적변수
custZip = ""
custZipVar = tk.StringVar()
custZipVar.set(custZip)

#위메프 수취인 주소 동적변수
wmpAddrRslt = ""
wmpAddrRsltVar = tk.StringVar()
wmpAddrRsltVar.set(wmpAddrRslt)

#위메프 우편번호 동적변수
wmpZipRsltt = ""
wmpZipRsltVar = tk.StringVar()
wmpZipRsltVar.set(wmpZipRsltt)

#작업완료여부 동적변수
jobDone = ""
jobDoneVar = tk.StringVar()
jobDoneVar.set(jobDone)

#주소라벨 동적변수
addrLb = "위메프주소:"
addrLbVar = tk.StringVar()
addrLbVar.set(addrLb)

#***********************************
def msgBox (msg) :
    window = Tk()
    window.wm_attributes("-topmost", 1)
    MsgBox = messagebox.showinfo("알림", msg)
    if(MsgBox == 'ok') :
        window.destroy()
    window.mainloop()

    
def varSet () :
        #기본동적변수 세팅
        global custNm
        custNm = load_ws.cell(xlRow, 27).value
        custNmVar.set(custNm)
        
        global wmpId
        wmpId = load_ws.cell(xlRow, 41).value
        wmpIdVar.set(wmpId)
        
        if(dicFlag) :
            wmpIdLb.delete(1.0,"end")
            wmpPwLb.delete(1.0,"end")
            if wmpId in idDic:
                wmpIdLb.insert(1.0, idDic[wmpId][0])
                wmpPwLb.insert(1.0, idDic[wmpId][1])
        global goodsNm
        goodsNm = load_ws.cell(xlRow, 13).value
        goodsNmVar.set(goodsNm)
        global goodsCnt
        goodsCnt = load_ws.cell(xlRow, 23).value + "개"
        goodsCntVar.set(goodsCnt)          
        global custAddr
        custAddr = load_ws.cell(xlRow, 30).value
        custAddrLb2.delete(1.0,"end")
        custAddrLb2.insert(1.0, custAddr)
        global custZip
        custZip = load_ws.cell(xlRow, 29).value
        custZipVar.set(custZip) 
        #우편번호 및 주소세팅
        global wmpAddrRsltVar
        global wmpZipRsltVar
        nvrAddrRslt = load_ws.cell(xlRow, maxCol+1).value
        wmpZipRslt = load_ws.cell(xlRow, maxCol+4).value
        flagYn = load_ws.cell(xlRow, maxCol+5).value
        
        if(wmpZipRslt != '' and wmpZipRslt is not None) :
            wmpAddrRsltVar.set(load_ws.cell(xlRow, maxCol+3).value)
            wmpZipRsltVar.set(load_ws.cell(xlRow, maxCol+4).value)
        elif(nvrAddrRslt != '' and nvrAddrRslt is not None and flagYn == 'Y') :
            wmpAddrRsltVar.set('위메프 주소입력 전')
            wmpZipRsltVar.set('')
        else :
            wmpAddrRsltVar.set('주소입력 오류(확인 필요) ')
            wmpZipRsltVar.set('')

        global jobDone
        jobDone = load_ws.cell(xlRow, maxCol+9).value
        if(jobDone is None) :
            jobDone = ""
        jobDoneVar.set(jobDone) 
        
def mkDic () :
    global idDic
    sheet_names = load_wb.sheetnames
    if(len(sheet_names) > 1) :
        global dicFlag 
        dicFlag = True
        second_sheet = sheet_names[1] 
        load_ws_dic = load_wb[second_sheet]
        maxRow_dic = load_ws_dic.max_row
        for i in range (2,maxRow_dic+1) :
            krId = load_ws_dic.cell(i, 1).value
            engId = load_ws_dic.cell(i, 2).value
            pw = load_ws_dic.cell(i, 3).value
            if(krId is not None and krId != '' and engId is not None and engId != '' and pw is not None and pw != '' ) :
                idDic[krId] = [engId, pw]
        
#다음버튼 누르면 xlROw 1증가하고 해당열의 수취인이름/ 로그인ID/ 쿠팡엑셀주소 / 우편번호 변경하기
def nextRow():
    if(xlDone) :
        global xlRow
        xlRow+=1
        rowVar.set(xlRow) 
        #동적변수 세팅
        varSet ()

    else :
        msgBox("엑셀작업을 먼저 수행하세요.")

#이전버튼 누르면 xlROw 1 감소하고 해당열의 수취인이름/ 로그인ID/ 쿠팡엑셀주소 / 우편번호 변경하기
def prevRow():
    if(xlDone) :
        global xlRow
        if(xlRow > 3) :
            xlRow-=1
            rowVar.set(xlRow) 
        #동적변수 세팅
        varSet ()
    else :
        msgBox("엑셀작업을 먼저 수행하세요.")    
def getFileName () :
    try : 
        global xlDone
        xlDone = False
        
        filePathObj = filedialog.askopenfile( title='파일 선택', filetypes=(('Excel files', '*.xlsx'), ('all files', '*.*'))) 
        filePathStr = str(filePathObj)
        idx1 = filePathStr.find("name=") + 5
        idx2 = filePathStr.find("mode=") - 1

        global filePath
        filePath = filePathStr[idx1:idx2].replace("'","")
        idx3 = filePath.rfind('/') + 1
        
        global fileName
        fileName = filePath[idx3:]
        fileNameVar.set(fileName)

        #파일 불러와서 네이버작업 된 애인지 확인하기. 되었으면 xlDone = True
        global load_wb 
        load_wb = load_workbook(filePath, data_only=True)
        #첫번째 시트 불러오기 
        sheet_names = load_wb.sheetnames
        first_sheet = sheet_names[0]
        global load_ws 
        load_ws = load_wb[first_sheet]        
        maxColTmp = load_ws.max_column #지역변수
        
        #기작업된 엑셀인지 확인
        for i in range (1,maxColTmp+1) : #1행에 있는 애들중에 조회
            if load_ws.cell(1,i).value == '네이버 주소검색 값' :
                xlDone = True
                #딕셔너리작업 및 동적변수 세팅
                mkDic ()
                global maxCol
                maxCol = i -1
                varSet()
                
        if(not xlDone) : #작업안된 엑셀이면
            global xlRow
            xlRow = 3
            global dicFlag
            dicFlag = False
            
    except Exception as e:
        msgBox (e)
        window.mainloop()
        pass

#************************************************************************************************************************************#
def doExel () :
    try :  
        global xlDone
        if(xlDone) :
            msgBox ("이미 작업완료된 엑셀입니다. 원본파일로 다시 시도하세요.")
            return
        global filePath
        if(filePath is None or filePath == '') :
            msgBox ("파일을 선택하세요.")
            return
        logger.info('엑셀시작')
        #백업파일 생성
        newFile = filePath[:-5] + "_working.xlsx"
        shutil.copy(filePath, newFile)
        filePath = newFile
        # data_only=True로 해줘야 수식이 아닌 값으로 받아온다.        
        global load_wb 
        load_wb = load_workbook(filePath, data_only=True)
        #첫번째 시트 불러오기 
        sheet_names = load_wb.sheetnames
        first_sheet = sheet_names[0]
        global load_ws 
        load_ws = load_wb[first_sheet]
        #행,열 크기 
        global maxRow 
        maxRow = load_ws.max_row
        
        global maxCol 
        
        maxCol = load_ws.max_column
        
        logger.info('전체 ' + str(maxRow) + '건 작업시작')
        
        #컬럼명 추가 
        nvrAddr = True  #네이버 주소검색 값 열추가 
        addrCls = True  #주소 구분(지번,도로명)
        wmpFullAddr = True #위메프 전체주소값
        wmpZipcd = True #위메프 우편번호	
        compYN  = True  #정상처리여부
        errTxt = True   #오류사유
        nvrRdXl = True #네이버도로명
        nvrAptXl = True #네이버 건물명
        #엑셀 작업하고 같은파일에 다시 엑셀작업할 경우
        

        for i in range (1,maxCol+1) : #1행에 있는 애들중에 조회
            if load_ws.cell(1,i).value == '네이버 주소검색 값' :
                maxCol -= 8
                nvrAddr = False
            if load_ws.cell(1,i).value == '주소구분' :
                addrCls = False
            if load_ws.cell(1,i).value == '위메프 전체주소값' :
                wmpFullAddr = False
            if load_ws.cell(1,i).value == '위메프 우편번호' :
                wmpZipcd = False
            if load_ws.cell(1,i).value == '정상처리여부' :
                compYN = False
            if load_ws.cell(1,i).value == '오류사유' :
                errTxt = False
            if load_ws.cell(1,i).value == '네이버도로명' :
                nvrRdXl = False
            if load_ws.cell(1,i).value == '네이버건물명' :
                nvrAptXl = False
                
        if nvrAddr : #열이 삽입되어 있지 않으면
            load_ws.cell(1,maxCol+1,'네이버 주소검색 값')
            
        if addrCls : #열이 삽입되어 있지 않으면
            load_ws.cell(1,maxCol+2,'주소구분')
            
        if wmpFullAddr : #열이 삽입되어 있지 않으면
            load_ws.cell(1,maxCol+3,'위메프 전체주소값')
            
        if wmpZipcd : #열이 삽입되어 있지 않으면
            load_ws.cell(1,maxCol+4,'위메프 우편번호')
            
        if compYN : #열이 삽입되어 있지 않으면
            load_ws.cell(1,maxCol+5,'정상처리여부')

        if errTxt : #열이 삽입되어 있지 않으면
            load_ws.cell(1,maxCol+6,'오류사유')
            
        if nvrRdXl : #열이 삽입되어 있지 않으면
            load_ws.cell(1,maxCol+7,'네이버도로명')

        if nvrAptXl : #열이 삽입되어 있지 않으면
            load_ws.cell(1,maxCol+8,'네이버건물명')           
        logger.info('엑셀 기초공사 종료')
        
    except Exception as e:
        logger.info(e)
        msgBox(e)
    finally :
        #저장하기    
        load_wb.save(filePath)
   
    logger.info('수취인 주소 네이버주소로 변경하기 START: ')
    #네이버에서 검색하기
    #[네이버] 엘리먼트
    homeSrchBoxEl = '//input[@class="input_text"]' #홈화면 검색창
    homeSrchBtnEl = '//button[@id="search_btn"]'   #홈화면 검색버튼 

    srchBoxEl = '//input[@id="nx_query"]'         #네이버 메인검색창
    srchBtnEl = '//button[@class="bt_search"]'    #네이버 메인검색버튼
    #네이버 홈화면 새창에서 열기
    options = webdriver.ChromeOptions() 
    options.add_argument("--disable-blink-features=AutomationControlled") 
    user_ag = UserAgent().random 
    options.add_argument('user-agent=%s'%user_ag) 
    options.add_experimental_option("excludeSwitches", ["enable-automation"]) 
    options.add_experimental_option("useAutomationExtension", False) 
    options.add_experimental_option("prefs", {"prfile.managed_default_content_setting.images": 2}) 

    #백그라운드 수행
    #options.add_argument("headless")    
    global driver 
    driver = webdriver.Chrome(drvPath, options=options) 

    # 크롤링 방지 설정을 undefined로 변경 

    driver.execute_cdp_cmd("Page.addScriptToEvaluateOnNewDocument", {
        "source": """ 
            Object.defineProperty(navigator, 'webdriver', { 
                get: () => undefined 
                }) 
                """ 
    }) 
    
    driver.get("https://www.naver.com/")

    time.sleep(SLEEPTIME * 2)
    
    #한번 임의 검색 때리고 다음창으로 이동해서 작업함.
    driver.find_element_by_xpath(homeSrchBoxEl).send_keys('a')
    driver.find_element_by_xpath(homeSrchBtnEl).click()

    time.sleep(SLEEPTIME * 2)

    logger.info('주소검색 반복문 START !')
    
    for k in range (2,maxRow+1) :
        try :
            #빈셀은 건너뛰기
            custName = load_ws.cell(k, 27).value 
            if(custName is None or custName == '') :
                continue
                
            logger.info( str(k) + "행 네이버검색 시작")

            #기존 입력값 지우기
            driver.find_element_by_xpath(srchBoxEl).clear()

            #수취인 주소검색
            rcvAddrExl = load_ws.cell(k, 30).value 
            driver.find_element_by_xpath(srchBoxEl).send_keys(rcvAddrExl)
            driver.find_element_by_xpath(srchBtnEl).click()
            
            time.sleep(SLEEPTIME)
            
            #주소결과값 가져오기
            addr01El = '//div[@class="ITiBH"]' 
            
            try :
                addr01 = driver.find_element_by_xpath(addr01El).text    #상세주소 제외한 주소
            except :
                continue
            #지번 vs 도로명 여부
            rdAddrYnEl = '//span[@class="LxiWh"]'
            try :
                rdYnTxt = driver.find_element_by_xpath(rdAddrYnEl).text
            except :
                continue
            #반대로 간다.
            if(rdYnTxt == '지번') : 
                rdYnTxt = '도로명'
            else : 
                rdYnTxt = '지번'
                try :
                    nvrRdAddr = driver.find_element_by_xpath('//*[@id="no-matched-address-list"]').text
                except :
                    nvrRdAddr = driver.find_element_by_xpath('//*[@id="unique"]').text
                    pass
                #네이버 전체주소
                try :
                    nvrFullAddr = driver.find_element_by_xpath('//*[@id="loc-main-section-root"]/section/div/div[2]/div[2]/div/div[1]/div[2]/div[1]/span[2]').text
                except : 
                    nvrFullAddr = "No result"
                    pass                
                if nvrRdAddr in nvrFullAddr : #단지명만 빼기
                    nvrAptNm = nvrFullAddr.replace("도로명","").split(nvrRdAddr)[0]
                    load_ws.cell(k,maxCol+7, nvrRdAddr)
                    load_ws.cell(k,maxCol+8, nvrAptNm)
  
            load_ws.cell(k,maxCol+1, addr01)
            load_ws.cell(k,maxCol+2, rdYnTxt)
            load_ws.cell(k,maxCol+5,'Y')
            
            #20건마다 저장
            if( k % 20 == 0 ) :
                load_wb.save(filePath) 
                
        except Exception as e:
            load_ws.cell(k,maxCol+5,'N')
            load_ws.cell(k,maxCol+6,'네이버주소검색 시 오류')
            load_ws.cell(k,maxCol+7, e)
            load_wb.save(filePath)
            msgBox(e)
            continue
        
        finally :
            load_wb.save(filePath)   #for문 끝나고 최종저장
    logger.info("엑셀 작업완료")

    xlDone = True
    #위메프 ID 딕셔너리 생성
    mkDic ()
    #동적변수 세팅
    varSet()
    #로그인 URL/ ID /PW
    driver.get(urlGlob)
    time.sleep(SLEEPTIME)
    msgBox("엑셀 작업완료")
    

#***엑셀끗****************************************************************************************************************************#
#상품페이지로 이동
def movePage () :
    try :
        tabs = driver.window_handles
        driver.switch_to.window(tabs[0])
        goodsUrl = load_ws.cell(xlRow,43).value
        driver.get(goodsUrl)
    except :
        msgBox (""+str(xlRow) + "행 상품페이지-URL 이상. 새창으로 다시 시도하세요")  
#***********
def focus() :
        tabs = driver.window_handles
        driver.switch_to.window(tabs[0])

#*****새창 ************************************************************************************************************#
def loginStart () :

    #로그인
    options = webdriver.ChromeOptions() 
    options.add_argument("--disable-blink-features=AutomationControlled") 
    user_ag = UserAgent().random 
    options.add_argument('user-agent=%s'%user_ag) 
    options.add_experimental_option("excludeSwitches", ["enable-automation"]) 
    options.add_experimental_option("useAutomationExtension", False) 
    options.add_experimental_option("prefs", {"prfile.managed_default_content_setting.images": 2}) 
   
    global driver 
    driver = webdriver.Chrome(drvPath, options=options) 

    # 크롤링 방지 설정을 undefined로 변경 

    driver.execute_cdp_cmd("Page.addScriptToEvaluateOnNewDocument", {
        "source": """ 
            Object.defineProperty(navigator, 'webdriver', { 
                get: () => undefined 
                }) 
                """ 
    }) 

    driver.get(urlGlob)
    
#*****신규배송지 입력*******************************************************************************************************************#
#사용자가 물품페이지로 이동하여 구매하기 버튼 누르고 난뒤의 작업(주소입력)
def inputNewAddr () :
    #신규배송지버튼
    newAddrEl = '//*[@id="newAddr"]'
    #신규배송지 클릭 
    driver.find_element_by_xpath(newAddrEl).click()

    #엑셀에서 값 받아오기
    #수취인명
    custName = load_ws.cell(xlRow,27).value
    #MOH 회사 폰번호
    custMobile2 = '7394'
    custMobile3 = '6512'
    #배송메시지
    shipMsg = load_ws.cell(xlRow,31).value

    #[위메프] 고객명 폰번호 우편번호찾기 엘리먼트
    custNameEl = '//input[@id="receiveName"]'
    custMobile2El = '//input[@id="mobile2"]'
    custMobile3El = '//input[@id="mobile3"]'
    selectMsgEl = '//a[@id="shipMsgDisplay"]'
    shipMsgEl = '//a[@data-shipmsg="직접 입력"]'
    shipMsgTxtEl = '//input[@id="directShipMsg"]'
    custZipcdEl = '//a[@class="btn_sys mid_d btn_zip"]'
    
    #수취인명 입력
    driver.find_element_by_xpath(custNameEl).send_keys(custName)
    #MOH 폰번호 중간 4자리 입력
    driver.find_element_by_xpath(custMobile2El).send_keys(custMobile2)
    #MOH 폰번호 뒷 4자리 입력
    driver.find_element_by_xpath(custMobile3El).send_keys(custMobile3)
    #배송시 요청사항 입력
    driver.find_element_by_xpath(selectMsgEl).click()
    driver.find_element_by_xpath(shipMsgEl).click()
    driver.find_element_by_xpath(shipMsgTxtEl).send_keys(shipMsg)
     
    #우편번호찾기 버튼 클릭
    driver.find_element_by_xpath(custZipcdEl).click()

    logger.info(str(xlRow) + "행(" + custName + ") 신규배송지 검색 시작")
    
#************************************************************************************************************************************#
#*****수취인 주소검색 및 입력************************************************************************************************************#    
#네이버에서 검색한 결과를 위메프 팝업창에 입력하여 조회. 
#위메프 조회결과중에 정확히 주소 찾기
def addrSearch () : 
    try :
        global xlRow
        #위메프 주소검색 팝업으로 이동
        tabs = driver.window_handles
        driver.switch_to.window(tabs[1])
    
        #수취인 주소 
        rcvAddrExl = load_ws.cell(xlRow, 30).value 
        logger.info( str(xlRow) + "행 위메프검색 시작")
        
        #지번 도로명 여부 
        rdYnExl = load_ws.cell(xlRow, maxCol +2).value
        
        #네이버 주소
        if(rdYnExl == '도로명') :
            nvrAddrExl = load_ws.cell(xlRow, maxCol +1).value 
        else : #입력값이 지번주소면 네이버도로명 세팅
            nvrAddrExl = load_ws.cell(xlRow, maxCol +7).value
        
        #상세주소 추출
        if(rdYnExl == '도로명') :
            cutter = len(nvrAddrExl)+1    #네이버 주소검색값(앞 주소) 
        else : #지번일 경우 다시 지번 세팅
            cutter = len(load_ws.cell(xlRow, maxCol +1).value)+1
        addrDtl = rcvAddrExl[cutter:] #상세주소
        
        #[위메프]주소검색 팝업창 엘리먼트
        custZipSrchEl = '//input[@title="검색어 입력"]'
        custZipCfmEl = '//a[@class="btns_sys red_mid_d"]'

        #주소입력
        driver.find_element_by_xpath(custZipSrchEl).clear()
        driver.find_element_by_xpath(custZipSrchEl).send_keys(nvrAddrExl)

        #조회버튼
        driver.find_element_by_xpath(custZipCfmEl).click()
        time.sleep(SLEEPTIME)

        #해당 창에 있는 주소검색 결과 수대로 작업
        rsltAddrEl = '//dl[@data-button="setAddress"]' #주소 결과값이 표시되는 엘리먼트
        rsltList = driver.find_elements_by_xpath(rsltAddrEl)
        rsltSize = len(rsltList)
        
        #검색성공 플래그
        findFlag = False 

        if(rsltSize == 1) : #조회결과가 1건이면 그냥 클릭
            findFlag = True
            buld = driver.find_element_by_xpath(rsltAddrEl)
        elif(rsltSize > 1) : #여러건일경우
            pageCntStr = driver.find_element_by_xpath('//a[@class="last"]').get_attribute("data-page")     #주소검색결과 전체 페이지수 
            pageCnt = int(pageCntStr)

            #페이지 하나씩 넘어가면서 주소찾기
            for j in range (1,pageCnt+1) :        
                pageEl = '//a[@data-page="' + str(j) + '"]'  #페이지 a태그 엘리먼트
                driver.find_element_by_xpath(pageEl).click() #페이지 클릭
                time.sleep(SLEEPTIME * 0.2)
                for i in range (rsltSize) :                    #화면에 있는 주소 갯수만큼 돌기
                    rsltList = driver.find_elements_by_xpath(rsltAddrEl)   #왜인지 모르겠지만 한번더 적어줘야함;
                    addrKeyword = nvrAddrExl                                          #네이버의 도로명주소
                    addrTarget  = rsltList[i].get_attribute("data-params-road-name1") #위메프의 도로명주소               
                    #검색결과중 원하는 주소 찾으면            엑셀주소에서 못찾으면 네이버 건물명에서 찾기
                    if(addrKeyword + ' ' in addrTarget or addrKeyword == addrTarget) : #찾는 주소가 위메프 주소안에 포함되면, 해당 엘리먼트 하위 태그 클릭
                        buld = rsltList[i]                              #찾는 주소 엘리먼트 
                        findFlag = True                                 #검색완료
                        break
                if(findFlag) : break
                
        #일치하는 주소 찾으면 엑셀에 값 넣고, 화면 클릭하기 
        if(findFlag) : 
            buldChildEl = './/a[@class="addr_v2 address"]'  #buld 엘리먼트는 클릭이 안돼서 하위 엘리먼트 클릭해야함. 
            
            #전체주소 값 엑셀에 입력 (위메프 주소검색결과+ 상세주소)
            wmpAddr = buld.find_element_by_xpath(buldChildEl).text
            
            #위메프 우편번호 엑셀에 입력
            zipCdEl = './/dt[@class="zipcode_v2"]' 
            zipCd = buld.find_element_by_xpath(zipCdEl).text
            load_ws.cell(xlRow, maxCol+4, zipCd)
            
            #쿠팡 vs 위메프 우편번호 비교
            global zipFlag
            zipFlag = True
            zipCpn = load_ws.cell(xlRow, 29).value
            if(zipCpn != zipCd) :
                load_ws.cell(xlRow,maxCol+5,'N')
                load_ws.cell(xlRow,maxCol+6,'우편번호 불일치')
                zipFlag = False
            else :
                #정상처리여부 Y 
                load_ws.cell(xlRow,maxCol+5,'Y')

            buld.find_element_by_xpath(buldChildEl).click() #하위 a태그 클릭
            
            time.sleep(SLEEPTIME * 0.3)
            
            driver.switch_to.window(tabs[0])
            #상세주소 입력
            buldNm = driver.find_element_by_xpath('//input[@id="addr2"]').get_attribute('value')
            if(buldNm != '' and buldNm is not None) :
                fbuldNm = wmpAddr.find(buldNm)
                wmpAddr = wmpAddr[:fbuldNm]
            load_ws.cell(xlRow, maxCol+3, wmpAddr + ' '+ addrDtl)

            driver.find_element_by_xpath('//input[@id="addr2"]').clear()
            driver.find_element_by_xpath('//input[@id="addr2"]').send_keys(addrDtl) 
        else :         #못찾았으면 창닫고 엑셀에는 오류기록 
            load_ws.cell(xlRow,maxCol+5,'N')
            load_ws.cell(xlRow,maxCol+6,'위메프 주소검색 결과없음')   
        #xlRow += 1               #한 열씩 증가. tkinter 화면버튼에서 구현
    except :
        load_ws.cell(xlRow,maxCol+5,'N')
        load_ws.cell(xlRow,maxCol+6,'위메프 주소검색 시 시스템오류')
        logger.info(str(xlRow) + '행 위메프 주소검색 시 시스템오류')
        pass
    finally :
        load_wb.save(filePath)  
        driver.switch_to.window(tabs[0]) #위메프 창으로 이동
    
def addrBtnOnclick () :    
    #주소 입력 클릭 시
    if(xlDone) :
        nvrAddrExl = load_ws.cell(xlRow, maxCol +1).value
        if( nvrAddrExl is None or nvrAddrExl == '') :
            load_ws.cell(xlRow,maxCol+5,'N')
            load_ws.cell(xlRow,maxCol+6,'네이버 또는 위메프 주소검색 결과 없음')
            load_wb.save(filePath)
            inputNewAddr()
            msgBox ("엑셀 " + str(xlRow) + "행 네이버 주소검색결과 없음")
        else :
            inputNewAddr ()
            time.sleep(SLEEPTIME)
            addrSearch () 

        #동적변수 세팅
        varSet ()
        
        if(not zipFlag) :           
            msgBox("우편번호 불일치")
            
    else :
        msgBox ("엑셀작업 완료 후 다시 시도하세요.")
        
def jobsDone () :
    if(xlDone) :
        load_ws.cell(xlRow,maxCol+9,'완료')
        #배경색 및 글자색 바꾸기
        y_color = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        b_font = Font(color = '000000')
        for i in range(1, maxCol+15):
            load_ws.cell(xlRow,i).fill = y_color
            load_ws.cell(xlRow,i).font = b_font
        load_wb.save(filePath)
        varSet ()
    else :
        msgBox("엑셀작업을 먼저 수행하세요.")
        
##****************************************************************************************
##****************************************************************************************
##****************************************************************************************
#11번가 함수 
#사용자가 물품페이지로 이동하여 구매하기 버튼 누르고 난뒤의 작업(주소입력)
def inputNewAddr_11 () :
    #신규배송지 클릭 
    driver.find_element_by_xpath('//*[@id="addrArrRadioRegLabel"]/span').click()

    #엑셀에서 값 받아오기
    #수취인명
    custName = load_ws.cell(xlRow,27).value
    #MOH 회사 폰번호
    custMobile2 = '7394'
    custMobile3 = '6512'
    #배송메시지
    shipMsg = load_ws.cell(xlRow,31).value

    #수취인명 입력
    driver.find_element_by_xpath('//*[@id="rcvrNm"]').send_keys(custName)
    #MOH 폰번호 중간 4자리 입력
    driver.find_element_by_xpath('//*[@id="tmprcvrPrtblNo2"]').send_keys(custMobile2)
    #MOH 폰번호 뒷 4자리 입력
    driver.find_element_by_xpath('//*[@id="tmprcvrPrtblNo3"]').send_keys(custMobile3)
    #배송시 요청사항 입력
    driver.find_element_by_xpath('//*[@id="ordDlvReqContSelect_ref"]').click()
    rsltList = driver.find_elements_by_xpath('//*[@id="ordDlvReqContSelect_ref"]/option')
    rsltSize = len(rsltList)
    
    for i in range (rsltSize) :                    
        rsltList = driver.find_elements_by_xpath('//*[@id="ordDlvReqContSelect_ref"]/option')                                         
        msgTarget  = rsltList[i].get_attribute("value")
        if(msgTarget == 'edit') : 
            rsltList[i].click() 
            break;
    
    driver.find_element_by_xpath('//*[@id="ordDlvReqContPrd_ref"]').clear()
    driver.find_element_by_xpath('//*[@id="ordDlvReqContPrd_ref"]').send_keys(shipMsg)
     
    #우편번호찾기 버튼 클릭
    driver.find_element_by_xpath('//*[@id="searchAddrBt"]').click()
    
#************************************************************************************************************************************#
#*****수취인 주소검색 및 입력************************************************************************************************************#    
#네이버에서 검색한 결과를 11번가 팝업창에 입력하여 조회. 
#11번가 조회결과중에 정확히 주소 찾기
def addrSearch_11 () : 
    try :
        global xlRow
        #11번가 주소검색 팝업으로 이동
        tabs = driver.window_handles
        driver.switch_to.window(tabs[1])
    
        #수취인 주소 
        rcvAddrExl = load_ws.cell(xlRow, 30).value 
        
        #지번 도로명 여부 
        rdYnExl = load_ws.cell(xlRow, maxCol +2).value
        
        #네이버 주소
        if(rdYnExl == '도로명') :
            nvrAddrExl = load_ws.cell(xlRow, maxCol +1).value 
        else : #입력값이 지번주소면 네이버도로명 세팅
            nvrAddrExl = load_ws.cell(xlRow, maxCol +7).value
        
        #상세주소 추출
        if(rdYnExl == '도로명') :
            cutter = len(nvrAddrExl)+1    #네이버 주소검색값(앞 주소) 
        else : #지번일 경우 다시 지번 세팅
            cutter = len(load_ws.cell(xlRow, maxCol +1).value)+1
        addrDtl = rcvAddrExl[cutter:] #상세주소
        
        #주소입력
        a =[]
        for i, c in enumerate(nvrAddrExl):
                if(c ==' ') :
                    a.append(i)

        if(len(a) < 3 ) : #세종과 같이 짧은 주소는 그냥 원래대로조회
            modfiedAddr = nvrAddrExl
        else :            #긴주소는 짜르기
            modfiedAddr = nvrAddrExl[a[1]+1:]
        
        driver.find_element_by_xpath('//*[@id="searchData"]').clear()
        driver.find_element_by_xpath('//*[@id="searchData"]').send_keys(modfiedAddr)

        #조회버튼
        driver.find_element_by_xpath('//button[@class="btn_search"]').click()
        time.sleep(SLEEPTIME)

        #해당 창에 있는 주소검색 결과 수대로 작업
        rsltSizeTxt = driver.find_element_by_xpath('//*[@id="totalSeachCnt"]').text
        if(rsltSizeTxt is None or rsltSizeTxt == '') :
            rsltSize = 0
        else :
            rsltSize = int(rsltSizeTxt)
        
        
        #검색성공 플래그
        findFlag = False 

        if(rsltSize == 1) : #조회결과가 1건이면 그냥 클릭
            findFlag = True
            buld = driver.find_element_by_xpath('//td[@class="addr"]')
        elif(rsltSize > 1) : #여러건일경우
            #스크롤 끝까지 내리기
            addr_panel = driver.find_element_by_xpath('//div[@class="list_box"]')
            scr_Cnt = int(rsltSize / 30)
            
            center_x = monitor_width/2 + 150
            center_y = monitor_height/2 + 20
            pg.click(x= center_x, y = center_y )
            if(rsltSize > 100) : 
                scr_Cnt = 1
            for j in range (scr_Cnt + 2) :
                time.sleep(SLEEPTIME)
                pg.scroll(-3000)

            for i in range (rsltSize) :                    #화면에 있는 주소 갯수만큼 돌기
                rsltList = driver.find_elements_by_xpath('//td[@class="addr"]')   #왜인지 모르겠지만 한번더 적어줘야함;
                addrKeyword = nvrAddrExl                                          #네이버의 도로명주소
                addrTarget  = rsltList[i].find_element_by_xpath('.//a').text           
                #검색결과중 원하는 주소 찾으면            엑셀주소에서 못찾으면 네이버 건물명에서 찾기
                if(addrKeyword + ' ' in addrTarget or addrKeyword == addrTarget) : #찾는 주소가 11번가 도로명안에 포함되면, 해당 엘리먼트 하위 태그 클릭
                    buld = rsltList[i] #찾는 주소 엘리먼트
                    findFlag = True                                  #검색완료
                    break
                
        #일치하는 주소 찾으면 엑셀에 값 넣고, 화면 클릭하기 
        if(findFlag) :             
            #11번가 검색결과 주소
            
            wmpAddr = buld.find_element_by_xpath('.//a').text  
            
            zipCd = buld.find_element_by_xpath('..').find_element_by_xpath('.//td[2]').text
            load_ws.cell(xlRow, maxCol+4, zipCd)
            
            #쿠팡 vs 11번가 우편번호 비교
            global zipFlag
            zipFlag = True
            zipCpn = load_ws.cell(xlRow, 29).value
            if(zipCpn != zipCd) :
                load_ws.cell(xlRow,maxCol+5,'N')
                load_ws.cell(xlRow,maxCol+6,'우편번호 불일치')
                zipFlag = False
            else :
                #정상처리여부 Y 
                load_ws.cell(xlRow,maxCol+5,'Y')
                
            buld.find_element_by_xpath('.//a').click() # a태그 클릭
            
            time.sleep(SLEEPTIME * 0.3)
            driver.switch_to.window(tabs[0])
            load_ws.cell(xlRow, maxCol+3, wmpAddr + ' '+ addrDtl)

            driver.find_element_by_xpath('//input[@id="rcvrDtlsAddr"]').clear()
            driver.find_element_by_xpath('//input[@id="rcvrDtlsAddr"]').send_keys(addrDtl) 
        else :         #못찾았으면 창닫고 엑셀에는 오류기록 
            driver.close()
            load_ws.cell(xlRow,maxCol+5,'N')
            load_ws.cell(xlRow,maxCol+6,'위메프 주소검색 결과없음')   
    except :
        driver.close()
        load_ws.cell(xlRow,maxCol+5,'N')
        load_ws.cell(xlRow,maxCol+6,'위메프 주소검색 시 시스템오류')
        pass
    finally :
        load_wb.save(filePath)  
        driver.switch_to.window(tabs[0]) #위메프 창으로 이동
        
#************************************************************************************************************************************#       
    
def addrBtnOnclick_11 () :    
    #주소 입력 클릭 시
    if(xlDone) :
        nvrAddrExl = load_ws.cell(xlRow, maxCol +1).value
        if( nvrAddrExl is None or nvrAddrExl == '') :
            load_ws.cell(xlRow,maxCol+5,'N')
            load_ws.cell(xlRow,maxCol+6,'네이버 또는 위메프 주소검색 결과 없음')
            load_wb.save(filePath)
            inputNewAddr_11()
            tabs = driver.window_handles
            if(len(tabs) > 1) :
                driver.switch_to.window(tabs[1])
                driver.close()
            tabs = driver.window_handles
            driver.switch_to.window(tabs[0])
            msgBox ("엑셀 " + str(xlRow) + "행 네이버 주소검색결과 없음")
        else :
            inputNewAddr_11 ()
            time.sleep(SLEEPTIME)
            addrSearch_11 () 

        #동적변수 세팅
        varSet ()
        
        if(not zipFlag) :           
            msgBox("우편번호 불일치")
            
    else :
        msgBox ("엑셀작업 완료 후 다시 시도하세요.")

def inputAddr () :
    if(whatTo == "위메프") :
        addrBtnOnclick()
    else :
        addrBtnOnclick_11()

def comboSet (event) :
    vari = combobox.get()
    global f1
    global urlGlob
    global whatTo
    if(vari == "위메프") :
        urlGlob = urlWmp
        addrLbVar.set("위메프주소:")
        whatTo = "위메프"
        msgBox("위메프 선택")
    else :
        urlGlob = url_11st
        addrLbVar.set("11번가주소:")
        whatTo = "11번가"
        msgBox("11번가 선택")

window.title("MOH x DHW AutoBoost")
window.wm_attributes("-topmost", 1)

rowNumLb = Label(window, textvariable=rowVar, width=3, height=3, font= fontStyle)
rowNumLb.place(x=4,y=5)

rowNumLb2 = Label(window, anchor = 'w', text="행 * ", width=4, height=3, font= fontStyle)
rowNumLb2.place(x=70,y=5)

custNmLb = Label(window, text=" 수취인 : ", width=7, height=3, font= fontStyle)
custNmLb.place(x=135,y=5)

custNmLb2 = Label(window, anchor = 'w', textvariable=custNmVar, wraplength =200, width=12, height=3, font= fontStyle)
custNmLb2.place(x=275,y=5)

goodsNmLb = Label(window, justify = 'center', textvariable= goodsNmVar ,anchor = 'nw', wraplength =200, width=22, height=5, font= tkFont.Font(family="Lucida Grande", size=13))
goodsNmLb.place(x=510,y=15)

goodsCntLb = Label(window, justify = 'center', textvariable= goodsCntVar ,anchor = 'nw', width=3, height=1, font= tkFont.Font(family="Lucida Grande", size=13, weight = "bold"))
goodsCntLb.place(x=470,y=50)

wmpIdLb = tk.Text(window, width=15, height=1.4 )
wmpIdLb.place(x=150,y=90)

wmpPwLb = tk.Text(window, width=15, height=1.4 )
wmpPwLb.place(x=280,y=90)

statusLb = Label(window, anchor = 'w', textvariable= addrLbVar, width=10, height=1, font= fontStyle2)
statusLb.place(x=10,y=120)

wmpAddrLb = Label(window, justify = 'left', textvariable= wmpAddrRsltVar ,anchor = 'nw', wraplength =550, width=55, height=3, font= fontStyle2)
wmpAddrLb.place(x=185,y=120)

wmpZipLb = Label(window, justify = 'left', textvariable= wmpZipRsltVar ,anchor = 'w', width=7, height=1, font= fontStyle2)
wmpZipLb.place(x=25,y=160)


custAddLb = Label(window, anchor = 'w', text= "쿠팡주소:", width=8, height=1, font=  fontStyle2)
custAddLb.place(x=15,y=220)


custAddrLb2 = tk.Text(window, width=45, height=3, font= fontStyle2 )
custAddrLb2.place(x=155,y=220)


custAddLb1 = Label(window, anchor = 'w', textvariable = custZipVar , width=7, height=1, font=  fontStyle2)
custAddLb1.place(x=25,y=260)
    

nextBtn = tk.Button(window, text="파일선택", command=getFileName, width=8, height=3)
nextBtn.place(x=15,y=355)

fileNameLb = Label(window, anchor = 'w', textvariable=fileNameVar,  wraplength =80, width=13, height=3)
fileNameLb.place(x=10,y=300)

nextBtn = tk.Button(window, text="엑셀작업", command=doExel, width=8, height=3)
nextBtn.place(x=115,y=355)

lineLb = Label(window, text="|", wraplength =1, width=1, height=1, anchor = 's', font= tkFont.Font(size=40) , fg = 'gray')
lineLb.place(x=205,y=355)


prevBtn = tk.Button(window, text="이전", command=prevRow, width=6, height=3, repeatdelay=10, repeatinterval=80)
prevBtn.place(x=255,y=355)

nextBtn = tk.Button(window, text="다음", command=nextRow, width=6, height=3, repeatdelay=10, repeatinterval=80)
nextBtn.place(x=325,y=355)


lineLb2 = Label(window, text="|", wraplength =1, width=1, height=1, anchor = 's', font= tkFont.Font(size=40) , fg = 'gray')
lineLb2.place(x=390,y=355)

toUrlBtn = tk.Button(window, text="상품페이지 이동", command=movePage, wraplength =80, width=8, height=3)
toUrlBtn.place(x=430,y=355)

typeAddrBtn = tk.Button(window, text="주소입력", command=inputAddr, width=10, height=3)
typeAddrBtn.place(x=515,y=355)

jobDoneBtn = tk.Button(window, text="작업완료", command=jobsDone, wraplength = 35, width=5, height=3)
jobDoneBtn.place(x=615,y=355)

newWinBtn = tk.Button(window, text="새창", command=loginStart, width=5, height=3)
newWinBtn.place(x=680,y=355)

jobDoneLb = Label(window, textvariable = jobDoneVar, width=3, height=2, font= fontStyle2)
jobDoneLb.place(x=610,y=300)

closeBtn = tk.Button(window, text="Focus", command=focus)
closeBtn.place(x=680,y=300)

combobox=tkinter.ttk.Combobox(window,  width=7, height=15, values= ["위메프","11번가"])
combobox.place(x=20,y=10)
combobox.bind("<<ComboboxSelected>>", comboSet)
combobox.set("위메프")


window.geometry('750x450')
window.resizable(False, False)
window.eval('tk::PlaceWindow . center')
window.mainloop()
